import os
import base64
from io import BytesIO
import re
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles.borders import BORDER_THIN, BORDER_HAIR, BORDER_MEDIUM
from openpyxl.utils import get_column_letter
from xlsx_validator import is_xlsx_base64, is_worksheet_valid

from cloudmersive import cloudmersive

# from getoutpdf import getoutpdf


def invalid_document_response():
    return {"error": "invalid document"}, 400


def dpformatter(file_base64):

    # Config
    ANLASS_ROW_HEIGHT = 350
    COLUMN_WIDTH = 10
    PAGE_MARGIN = 0.39  # 1 cm

    if not is_xlsx_base64(file_base64):
        return invalid_document_response()

    xlsx_base64 = re.sub(
        "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,",
        "",
        file_base64,
    )
    xlsxFile = BytesIO(base64.b64decode(xlsx_base64))
    wb = Workbook()
    wb = load_workbook(xlsxFile)
    ws = wb.worksheets[0]

    if not is_worksheet_valid(ws):
        return invalid_document_response()

    anlass_rows = []
    columns_to_fix_width_range = []  # Column range between Anlass and Kommentar
    vertical_text_cells_ranges = []
    kommentar_text_cells = []

    # Set page orientation, paper size and margins
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.top = PAGE_MARGIN
    ws.page_margins.left = PAGE_MARGIN
    ws.page_margins.right = PAGE_MARGIN
    ws.page_margins.bottom = PAGE_MARGIN
    ws.page_margins.header = PAGE_MARGIN
    ws.page_margins.footer = PAGE_MARGIN

    dp_index = 0
    for row in ws.iter_rows():
        for cell in row:
            if "Kommentar" in str(cell.value):
                columns_to_fix_width_range.append(cell.column - 1)
                kommentar_text_cells.append(
                    f"{get_column_letter(cell.column)}{cell.row+5}"
                )
                vertical_text_cells_ranges.append(
                    {"end": {"row": cell.row, "column": cell.column - 1}}
                )
            if "Anlass" in str(cell.value):
                columns_to_fix_width_range.append(cell.column + 2)
                anlass_rows.append(cell.row)
                vertical_text_cells_ranges[dp_index] = {
                    **vertical_text_cells_ranges[dp_index],
                    "start": {"row": cell.row, "column": cell.column + 2},
                }
                dp_index += 1

            # Fix BORDER_HAIR
            if cell.border.top.style == BORDER_HAIR:
                cell.border.top.style = BORDER_THIN
            if cell.border.bottom.style == BORDER_HAIR:
                cell.border.bottom.style = BORDER_THIN
            if cell.border.left.style == BORDER_HAIR:
                cell.border.left.style = BORDER_THIN
            # Right border style of one cell is sometimes
            # different than left border style of adjacent cell.
            # In this specific case: when right border style is HAIR
            # it should actually be MEDIUM
            if cell.border.right.style == BORDER_HAIR:
                cell.border.right.style = BORDER_MEDIUM

    # Set vertical text and wrap
    for vertical_text_cell_range in vertical_text_cells_ranges:
        for vertical_text_cell_column in range(
            vertical_text_cell_range["start"]["column"],
            vertical_text_cell_range["end"]["column"] + 1,
        ):
            ws.cell(
                vertical_text_cell_range["start"]["row"], vertical_text_cell_column
            ).alignment = Alignment(text_rotation=90, wrap_text=True)
            ws.cell(
                vertical_text_cell_range["start"]["row"] + 1, vertical_text_cell_column
            ).alignment = Alignment(text_rotation=90, wrap_text=True)

    # Wrap Kommentar Text
    for kommentar_text_cell in kommentar_text_cells:
        ws[kommentar_text_cell].alignment = Alignment(wrap_text=True, vertical="center")

    # Set Anlass rows height
    for anlass_row in anlass_rows:
        ws.row_dimensions[anlass_row].height = ANLASS_ROW_HEIGHT

    # Set width of columns between Anlass and Kommentar
    columns_to_fix_width_range.sort()
    coulmns_to_fix_width_range = range(
        columns_to_fix_width_range[0],
        columns_to_fix_width_range[len(columns_to_fix_width_range) - 1],
    )
    for coulmn_to_fix_width in coulmns_to_fix_width_range:
        ws.column_dimensions[
            get_column_letter(coulmn_to_fix_width)
        ].width = COLUMN_WIDTH

    # Set print layout
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    xlsx_output = BytesIO()
    wb.save(xlsx_output)

    wb.close()
    xlsx_output_base64 = base64.encodebytes(xlsx_output.getvalue()).decode("UTF-8")

    # Generate PDF via cloudmersive api
    with open("temp.xlsx", "wb") as f:
        f.write(xlsx_output.getbuffer())
    pdf_base64 = cloudmersive("temp.xlsx")

    # Delete temp.xlsx
    if os.path.isfile("temp.xlsx"):
        os.remove("temp.xlsx")

    # Generate PDF via getoutpdf.com api
    # pdf_base64 = getoutpdf(xlsx_output_base64)

    return {
        "pdfBase64": "data:application/pdf;base64," + pdf_base64,
        "xlsxBase64": "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,"
        + xlsx_output_base64,
    }
